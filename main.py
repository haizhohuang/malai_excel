from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill
from dataclasses import dataclass, field
from typing import List, Dict, Tuple
from copy import copy

from openpyxl.utils import get_column_letter

# 样式常量
BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)


@dataclass
class CellStyle:
    """单元格样式配置"""
    border: Border = BORDER
    fill: PatternFill = None


@dataclass
class DataGrid:
    """自包含的数据格对象"""
    start_row: int
    start_col: int
    parent: 'DataGrid' = None
    children: List['DataGrid'] = field(default_factory=list)
    style: CellStyle = CellStyle()

    def __post_init__(self):
        """初始化时生成单元格数据和连接路径"""
        self._cells = self._generate_cells()
        self.connection_path = self._calculate_connection_path()

    @property
    def end_row(self) -> int:
        """结束行号"""
        return self.start_row + 1  # 固定2行

    @property
    def end_col(self) -> int:
        """结束列号"""
        return self.start_col + 2  # 固定3列

    def _generate_cells(self) -> Dict[Tuple[int, int], str]:
        """生成单元格数据字典"""
        return {
            (r, c): f"{r}-{c}"
            for r in range(self.start_row, self.end_row + 1)
            for c in range(self.start_col, self.end_col + 1)
        }

    def _calculate_connection_path(self) -> Dict[Tuple[int, int], str]:
        """计算连接线路径（根据新规则）"""
        if not self.parent:
            return {}

        path = {}
        # 水平段路径（父结束列+1 到 子开始列-1）
        if self.start_col > self.parent.end_col:
            for col in range(self.parent.end_col, self.start_col + 1):
                coord = (self.parent.start_row, col)
                path[coord] = 'horizontal'

        # 垂直段路径（父结束行+1 到 子开始行-1）
        if self.start_row > self.parent.end_row:
            for row in range(self.parent.end_row, self.start_row + 1):
                coord = (row, self.start_col - 1)
                path[coord] = 'vertical'

        # 交汇点处理（最后一个水平段单元格）
        if self.start_row > self.parent.end_row:
            last_h_coord = (self.start_row, self.start_col - 1)
            if last_h_coord in path:
                path[last_h_coord] = 'corner'

        return path

    def get_cell_data(self, row: int, col: int) -> str:
        """获取单元格数据"""
        return self._cells.get((row, col), "")


class GridFactory:
    """数据格工厂类（根据新位置规则）"""

    @staticmethod
    def create_sample_grids() -> DataGrid:
        """创建符合示例需求的结构"""
        # 根数据格（A1-C2）
        root = DataGrid(1, 1, style=CellStyle(
            fill=PatternFill(start_color="FFA500", fill_type="solid")
        ))

        # 第一个子数据格（G1-I2）
        child1 = GridFactory.create_child(root, horizontal=True)

        # 第二个子数据格（G5-I6）
        child2 = GridFactory.create_child(root, horizontal=False)

        # 孙子数据格（M1-O2）
        GridFactory.create_child(child1, horizontal=True)

        GridFactory.create_child(child1, horizontal=False)

        return root

    @staticmethod
    def create_child(parent: DataGrid, horizontal: bool) -> DataGrid:
        """创建子数据格"""
        if horizontal:
            # 水平方向：右侧3列（间隔3列）
            start_row = parent.start_row
            start_col = parent.end_col + 4  # 3列内容 + 1列间隔
        else:
            # 垂直方向：下方2行（间隔2行）
            start_row = parent.end_row + 3  # 2行内容 + 1行间隔
            start_col = parent.end_col + 4

        child = DataGrid(
            start_row=start_row,
            start_col=start_col,
            parent=parent,
            style=CellStyle(fill=PatternFill(start_color="0000FF", fill_type="solid"))
        )
        parent.children.append(child)
        return child


class ExcelGenerator:
    """Excel生成器（支持新连接线规则）"""

    def __init__(self, filename: str):
        self.filename = filename
        self.wb = Workbook()
        self.ws = self.wb.active
        self.root = GridFactory.create_sample_grids()

    def _write_grid(self, grid: DataGrid):
        """写入数据格内容"""
        for row in range(grid.start_row, grid.end_row + 1):
            for col in range(grid.start_col, grid.end_col + 1):
                cell = self.ws.cell(row=row, column=col)
                cell.value = grid.get_cell_data(row, col)
                cell.border = grid.style.border
                if grid.style.fill:
                    cell.fill = grid.style.fill
                self._adjust_column_width(col)

    def _draw_connections(self, grid: DataGrid):
        """递归绘制连接线"""
        for coord, line_type in grid.connection_path.items():
            cell = self.ws.cell(row=coord[0], column=coord[1])
            self._apply_connection_style(cell, line_type)

        for child in grid.children:
            self._draw_connections(child)

    def _apply_connection_style(self, cell, line_type: str):
        """应用连接线样式"""
        new_border = copy(cell.border)

        if line_type == 'horizontal':
            new_border.bottom = BORDER.bottom
        elif line_type == 'vertical':
            new_border.left = BORDER.left
        elif line_type == 'corner':
            new_border.left = BORDER.left
            new_border.bottom = BORDER.bottom

        cell.border = new_border

    def _adjust_column_width(self, col: int):
        """统一调整列宽"""
        self.ws.column_dimensions[get_column_letter(col)].width = 12

    def generate(self):
        """生成完整文件"""
        # 写入所有数据格
        stack = [self.root]
        while stack:
            current = stack.pop()
            self._write_grid(current)
            stack.extend(reversed(current.children))

        # 绘制所有连接线
        self._draw_connections(self.root)

        self.wb.save(self.filename)


if __name__ == "__main__":
    generator = ExcelGenerator("modified_connection.xlsx")
    generator.generate()