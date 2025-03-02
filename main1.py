from copy import copy

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill
from dataclasses import dataclass
from typing import List, Tuple

# 样式常量
BORDER = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
ORANGE_FILL = PatternFill(start_color="FFA500", fill_type="solid")
BLUE_FILL = PatternFill(start_color="0000FF", fill_type="solid")


@dataclass
class GridStyle:
    """数据格样式配置"""
    border: Border = BORDER
    fill: PatternFill = None


@dataclass
class DataGrid:
    """层级数据格对象"""
    start_row: int
    start_col: int
    parent: 'DataGrid' = None
    children: List['DataGrid'] = None
    style: GridStyle = GridStyle()

    @property
    def range(self) -> Tuple[Tuple[int, int], Tuple[int, int]]:
        """获取数据格占据范围"""
        return (
            (self.start_row, self.start_row + 1),  # 2行
            (self.start_col, self.start_col + 2)  # 3列
        )

    @property
    def connection_path(self) -> List[Tuple[int, int]]:
        """计算连接线路径坐标"""
        if not self.parent:
            return []

        # 获取父节点和当前节点的中心坐标
        (p_row1, p_row2), (p_col1, p_col2) = self.parent.range
        (c_row1, c_row2), (c_col1, c_col2) = self.range

        # 生成连接路径（直角折线）
        path = []
        # 水平延伸段
        if p_col2 < c_col1:
            for col in range(p_col2 + 1, c_col1 + 1):
                path.append((p_row1, col))
        # 垂直延伸段
        if p_row2 < c_row1:
            for row in range(p_row2, c_row1):
                path.append((row, c_col1 - 1))
        return path


"""层级数据格生成器"""
class GridGenerator:

    """创建示例数据结构"""
    @staticmethod
    def create_sample_grids() -> DataGrid:
        root = DataGrid(1, 1, style=GridStyle(fill=ORANGE_FILL))

        # 第一层子节点
        child1 = DataGrid(1, 7, parent=root, style=GridStyle(fill=ORANGE_FILL))
        child2 = DataGrid(5, 7, parent=root, style=GridStyle(fill=BLUE_FILL))

        # 第二层子节点
        grandchild1 = DataGrid(1, 13, parent=child1, style=GridStyle(fill=BLUE_FILL))

        # 构建树结构
        root.children = [child1, child2]
        child1.children = [grandchild1]
        return root


"""Excel 渲染器"""
class ExcelRenderer:

    def __init__(self, filename: str):
        self.filename = filename
        self.wb = Workbook()
        self.ws = self.wb.active
        self.grid_root = GridGenerator.create_sample_grids()

    """渲染单个数据格"""
    def _render_grid(self, grid: DataGrid):
        (row_start, row_end), (col_start, col_end) = grid.range

        # 写入数据并应用基础样式
        for row in range(row_start, row_end + 1):
            for col in range(col_start, col_end + 1):
                cell = self.ws.cell(row=row, column=col)
                cell.value = f"{row}-{col}"
                cell.border = grid.style.border
                if grid.style.fill:
                    cell.fill = grid.style.fill

                # 自动调整列宽
                self.ws.column_dimensions[cell.column_letter].width = 15

    """递归绘制连接线"""
    def _draw_connections(self, grid: DataGrid):
        for child in grid.children or []:
            # 绘制连接路径
            for (row, col) in child.connection_path:
                cell = self.ws.cell(row=row, column=col)

                # 合并水平和垂直方向的边框
                new_border = copy(cell.border)
                if row == child.connection_path[0][0]:  # 垂直段
                    new_border.bottom = BORDER.bottom
                elif row == child.range[0][0] - 1:
                    new_border.left = BORDER.left
                    new_border.bottom = BORDER.bottom
                else:  # 水平段
                    new_border.left = BORDER.left
                cell.border = new_border

            # 递归处理子节点
            self._draw_connections(child)

    """生成完整表格"""
    def generate(self):
        # 先渲染所有数据格
        grids = []
        stack = [self.grid_root]
        while stack:
            current = stack.pop()
            grids.append(current)
            stack.extend(reversed(current.children or []))

        for grid in grids:
            self._render_grid(grid)

        # 绘制所有连接线
        self._draw_connections(self.grid_root)

        self.wb.save(self.filename)


# 使用示例
if __name__ == "__main__":
    renderer = ExcelRenderer("hierarchical_grids_with_connections.xlsx")
    renderer.generate()